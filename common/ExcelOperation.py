# -*- coding: utf-8 -*-
# @Time : 2021/11/25 15:47
# @Author : zhangyan
# @File : ExcelOperation.py  excel表相关操作
import os

import openpyxl


class ExcelOperation(object):
    # 初始化方法
    def __init__(self, filename):
        # 动态获取文件路径
        self.file_path = os.path.join(os.path.dirname(os.getcwd()), "data", filename)
        # 打开excel文件，返回workbook对象（用于获取sheet工作表）
        self.wb = openpyxl.load_workbook(self.file_path)

    # 读取excel文件方法
    def read_excel(self):
        pass

    # 写入excel文件方法，传入写入sheet表，行与列（列表形式），数值
    def write_excel(self, sheet_name, x_y, content):
        try:
            # 获取对应工作表
            sheet = self.wb[sheet_name]
            # 写入行与列
            sheet.cell(x_y[0], x_y[1]).value = content
        except Exception as e:
            sheet.cell(x_y[0], x_y[1]).value = content
        finally:
            # 保存文件
            self.wb.save(self.file_path)


if __name__ == '__main__':
    filename = "5GClient_testcase.xlsx"
    ExcelOperation(filename).write_excel("登录", [2, 13], "cccccccc")
