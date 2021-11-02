# -*- coding: utf-8 -*-
# @Time : 2021/11/1 15:48
# @Author : zhangyan
# @File : read_testcase.py  工具类：读取测试用例
import json
import os
import openpyxl


class ReadTestCase(object):
    # 初始化方法
    def __init__(self, filename):
        # 动态获取文件路径
        # print(os.path.abspath(__file__))  # 获取当前文件的绝对路径
        # print(os.path.dirname(os.path.abspath(__file__)))   # 获取当前文件的上一级目录的绝对路径
        self.filename = filename
        self.file_path = os.path.dirname(
            os.path.dirname(os.path.abspath(__file__))) + os.sep + "data" + os.sep + self.filename
        print(self.file_path)

        # 打开excel文件，返回一个workbook对象（用于获取sheet工作表）
        self.wb = openpyxl.load_workbook(self.file_path)

        # 返回所有工作表名列表['登录','总体态势',...]
        self.sheets = self.wb.sheetnames

        # 循环获取sheet表单对象
        self.sheet = self.wb[self.sheets[1]]

        self.rows = self.sheet.max_row  # 获取总行数（读取数据，遍历数据使用）
        print("总行数为：", self.rows)
        # self.columns = self.sheet.max_column  # 获取总列数
        # print("总列数为：", self.columns)

        # excel数据对应列
        self.cell_config = {
            "case_id": 1,
            "api_name": 2,
            "case_desc": 3,
            "path": 4,
            "method": 5,
            "headers": 6,
            "param_type": 7,
            "params": 8,
            "sql": 9,
            "except": 10,
            "is_run": 11,
            "result": 12,
            "desc": 13
        }

    # 读取excel
    def read_excel(self):
        # 新建空列表，存储每行测试用例
        caseList = list()

        # 遍历excel某表单每行数据
        for i in range(2, self.rows + 1):
            # 新建空字典，存储具体数据
            caseDict = dict()
            # 判断是否执行
            if self.sheet.cell(i, self.cell_config.get("is_run")).value == '是':
                try:
                    # 读取数据，追加到字典
                    caseDict['case_id'] = self.sheet.cell(i, self.cell_config.get("case_id")).value
                    caseDict['api_name'] = self.sheet.cell(i, self.cell_config.get("api_name")).value
                    caseDict['case_desc'] = self.sheet.cell(i, self.cell_config.get("case_desc")).value
                    caseDict['path'] = self.sheet.cell(i, self.cell_config.get("path")).value
                    caseDict['method'] = self.sheet.cell(i, self.cell_config.get("method")).value
                    # caseDict['headers'] = eval(self.sheet.cell(i, self.cell_config.get("headers")).value)
                    caseDict['headers'] = self.sheet.cell(i, self.cell_config.get("headers")).value
                    caseDict['param_type'] = self.sheet.cell(i, self.cell_config.get("param_type")).value
                    # caseDict['params'] = eval(self.sheet.cell(i, self.cell_config.get("params")).value)
                    caseDict['params'] = self.sheet.cell(i, self.cell_config.get("params")).value
                    caseDict['sql'] = self.sheet.cell(i, self.cell_config.get("sql")).value
                    # caseDict['except'] = eval(self.sheet.cell(i, self.cell_config.get("except")).value)
                    caseDict['except'] = self.sheet.cell(i, self.cell_config.get("except")).value

                    # 将字典追加到列表中
                    caseList.append(caseDict)

                    # 若读取成功，将读取结果写入excel中
                    self.write_excel([i, self.cell_config.get("desc")], "数据读取完成~")
                except Exception as e:
                    self.write_excel([i, self.cell_config.get("desc")], e)
        # 将列表数据写入json
        self.write_json("5GCenter_testCase.json",caseList)

    # 写入excel，依赖写入的行、列和数值
    def write_excel(self, x_y, content):
        try:
            # x_y参数格式为列表：[2,13]
            self.sheet.cell(x_y[0], x_y[1]).value = content
        except Exception as e:
            self.sheet.cell(x_y[0], x_y[1]).value = e
        finally:
            # 写入数据后，需要保存excel
            self.wb.save(self.file_path)

    # 读取json
    # def read_json(self):
    #     pass

    # 写入Json
    def write_json(self, filename, data):
        filepath = os.path.dirname(
            os.path.dirname(os.path.abspath(__file__))) + os.sep + "data" + os.sep + filename
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)


if __name__ == '__main__':
    # ReadTestCase("testcase.xlsx").write_excel([2, 13], "写入成功~")
    case_list = ReadTestCase("5GCenter_testcase.xlsx").read_excel()
    # print(testcaseData)
    ReadTestCase("5GCenter_testcase.json").write_json()