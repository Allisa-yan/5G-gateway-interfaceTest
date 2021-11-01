# -*- coding: utf-8 -*-
# @Time : 2021/11/1 15:44
# @Author : zhangyan
# @File : excelOperation.py   封装所有excel操作方法
import os

import openpyxl


class ExcelOperation(object):
    # 初始化
    def __init__(self, filename):
        # 动态获取文件路径
        self.file_path = os.path.dirname(__file__) + os.sep + "data" + os.sep + filename
        # 打开文件,通过openpyxl类获取文件对象workbook
        self.wb = openpyxl.load_workbook(self.filename)
        # 获取sheet工作表单对象
        self.sheet = self.wb[self.wb.sheetnames[0]]
        # 获取总行数（读取数据，遍历数据使用）
        self.rows = self.sheet.max_row
        print("总行数为：", self.rows)

    # 读取excel
    def read_excel(self):
        # 新建空列表，存储每行数据

        # 遍历每行数据
            # 新建空字典，存储每行数据

            # 判断是否执行
                # 读取数据，追加到字典

                # 将字典追加到列表中

                # 将读取结果写入excel中

        # 将列表数据写入json
        return

    # 写入excel
    def write_excel(self):
        pass
