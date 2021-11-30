# -*- coding: utf-8 -*-
# @Time : 2021/11/25 15:47
# @Author : zhangyan
# @File : excel_operation.py  excel表相关操作
import os

import openpyxl


class ExcelOperation(object):
    # 初始化方法
    def __init__(self, filename):
        # 动态获取文件路径
        self.file_path = os.path.join(os.path.dirname(os.getcwd()), "data", filename)
        # 打开excel文件，返回workbook对象（用于获取sheet工作表）
        self.wb = openpyxl.load_workbook(self.file_path)

    # 读取excel文件方法
    def read_excel(self):
        # 新建总的空字典，存储每张sheet表的所有测试用例
        caseDictTol = dict()
        # 获取所有sheet表名列表['登录','总体态势',...]
        sheets = self.wb.sheetnames
        print("所有sheet表名：", sheets)
        # # 接口测试用例每张sheet表的格式应一致，因此获取其中某个表单对象的总列数及对应字段
        # cols = self.wb[sheets[0]].max_column
        # print("最大列数：", cols)
        # 遍历所有sheet表
        for sheetname in sheets:
            # 获取当前sheet表总行数
            sheet = self.wb[sheetname]
            rows = sheet.max_row
            # 新建每张sheet表空列表，用于存放所有测试用例
            caseList = list()
            # 遍历excel某sheet表中每行数据
            for i in range(2, rows + 1):
                # 新建空字典，存储具体数据
                caseDict = dict()
                try:
                    # 读取数据，追加到字典中
                    caseDict['id'] = sheet.cell(i, 1).value
                    caseDict['c_name'] = sheet.cell(i, 2).value
                    caseDict['test_module'] = sheet.cell(i, 3).value
                    caseDict['project_id'] = sheet.cell(i, 4).value
                    caseDict['api_name'] = sheet.cell(i, 5).value
                    caseDict['api_path'] = sheet.cell(i, 6).value
                    caseDict['request_type'] = sheet.cell(i, 7).value
                    caseDict['headers'] = sheet.cell(i, 8).value
                    caseDict['params_type'] = sheet.cell(i, 9).value
                    caseDict['params'] = sheet.cell(i, 10).value
                    caseDict['sql'] = sheet.cell(i, 11).value
                    caseDict['expectation'] = sheet.cell(i, 12).value
                    caseDict['is_run'] = sheet.cell(i, 13).value
                    caseDict['create_time'] = sheet.cell(i, 14).value
                    caseDict['update_time'] = sheet.cell(i, 15).value
                    caseDict['user_id'] = sheet.cell(i, 15).value
                    caseDict['response'] = sheet.cell(i, 15).value
                    # 将每行测试用例数据，追加到该sheet表的总用例列表中
                    caseList.append(caseDict)
                except Exception as e:
                    print(e)
            # 将字典追加到总字典中
            caseDictTol[sheetname] = caseList
        print("最终读取excel结果为：", caseDictTol)

    # 写入excel文件方法，传入写入sheet表，行与列（列表形式），数值
    def write_excel(self, sheet_name, x_y, content):
        try:
            # 获取对应工作表
            sheet = self.wb[sheet_name]
            # 写入行与列
            sheet.cell(x_y[0], x_y[1]).value = content
        except Exception as e:
            sheet.cell(x_y[0], x_y[1]).value = content
        finally:
            # 保存文件
            self.wb.save(self.file_path)


if __name__ == '__main__':
    filename = "5GClient_testcase.xlsx"
    # ExcelOperation(filename).write_excel("登录", [2, 13], "cwwwwc")
    ExcelOperation(filename).read_excel()
